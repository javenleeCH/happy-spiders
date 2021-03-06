# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html

from openpyxl import load_workbook
from scrapy.pipelines.files import FilesPipeline
import os
import scrapy
import re
import sys


data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '../data')


class DownlodImagePipeline(FilesPipeline):
    def get_media_requests(self, item, info):
        year = item['time']
        images = item['images']
        name = item['name']
        i = 1
        for img_url in images:
            if 'http' in img_url:
                image_name = name + '_' + str(i)
                i += 1
                yield scrapy.Request(url=img_url, meta={'image_name': image_name, 'year': year})

    def file_path(self, request, response=None, info=None):
        image_name = request.meta['image_name']
        image_name = re.sub(r'/', ' ', image_name)
        year = request.meta['year']
        path = '%s/%s.jpg' % (str(year), image_name)
        return path


class DesignPipeline(object):
    def __init__(self):
        self.file = os.path.join(data_dir, '2017.xlsx')
        self.excel = load_workbook(self.file)
        self.ws = self.excel.active

    def process_item(self, item, spider):
        # self.file = os.path.join(data_dir, '1954.xlsx')
        # self.excel = load_workbook(self.file)
        # self.ws = self.excel.active
        name = item['name']
        type = item['type']
        discipline = item['discipline']
        year = item['year']
        development = item['development']
        regions = item['regions']
        groups = item['groups']
        criteria = item['criteria']
        clients = item['clients']
        universities = item['universities']
        designs = item['designs']
        images = item['images']
        description = item['description']

        clients_length = len(clients)
        if clients_length > 5:
            clients = clients[0:5]
        all_clients = []
        for client in clients:
            all_clients.append(client['manufacturer'])
            all_clients.append(client['location'])
        for i in range(5-clients_length):
            all_clients.append('')
            all_clients.append('')

        universities_length = len(universities)
        all_universities = []
        for universitie in universities:
            all_universities.append(universitie['school'])
            all_universities.append(universitie['location'])
        for i in range(5 - universities_length):
            all_universities.append('')
            all_universities.append('')

        designs_length = len(designs)
        if designs_length > 5:
            designs = designs[0:5]
        all_designs = []
        for design in designs:
            all_designs.append(design['designer'])
            all_designs.append(design['design_company'])
            all_designs.append(design['location'])
        for i in range(5 - designs_length):
            all_designs.append('')
            all_designs.append('')
            all_designs.append('')

        images_length = len(images)
        if images_length > 7:
            images = images[0:7]
        for i in range(7 - images_length):
            images.append('')

        client_uni_design_images = all_clients + all_universities + all_designs + images

        all_data = [name, type, discipline, year, development, regions, groups, criteria] + client_uni_design_images
        all_data.append(description)
        self.ws.append(all_data)
        self.excel.save(self.file)
        return item

    def close_spider(self, spider):
        self.excel.close()
